#!/usr/bin/env python3
"""
Audico EC → Peachtree Quote CSV Converter
==========================================

Reads an EC_XX_YYYY.xlsx monthly statement and produces one Peachtree-ready
Sage 50 Sales Journal CSV per yellow-highlighted event, plus summary & audit files.

Usage:
    python convert.py <EC_xlsx> [--config config/hotel_mapping.json] [--out ./out] [--customer-list LISTA.xlsx]

Design:
- Yellow-highlighted CLIENTE rows = items that need a quote.
- Each quote = 1 ITBMS tax line + N product lines + '***' separator + Evento line + Fecha line
- Line amounts are NEGATIVE (Peachtree sales credit convention). AR amount is POSITIVE total.
- Unit price = AUDICO U (column I) which is 50% of customer cost. We invoice the hotel at the 50%.
"""

import argparse
import csv
import json
import re
import sys
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

# Sage 50 Pro import schema — 59 columns, exact order, NO header row in the output file.
# Reverse-engineered from the actual Sage 50 Pro Sales Journal export at Audico, verified column-by-column.
# Fields like Customer Name, Accounts Receivable Amount, AR Date Cleared, Inventory Account, etc.
# are NOT included — Sage computes those automatically from the Customer ID and amounts.
SALES_HEADER = [
    "Customer ID",                     # 1
    "Invoice/CM #",                    # 2
    "Apply to Invoice Number",         # 3
    "Credit Memo",                     # 4
    "Progress Billing Invoice",        # 5
    "Date",                            # 6
    "Ship By",                         # 7
    "Quote",                           # 8
    "Quote #",                         # 9
    "Quote Good Thru Date",            # 10
    "Drop Ship",                       # 11
    "Ship to Name",                    # 12
    "Ship to Address-Line One",        # 13
    "Ship to Address-Line Two",        # 14
    "Ship to City",                    # 15
    "Ship to State",                   # 16
    "Ship to Zipcode",                 # 17
    "Ship to Country",                 # 18
    "Customer PO",                     # 19
    "Ship Via",                        # 20
    "Ship Date",                       # 21
    "Date Due",                        # 22
    "Discount Amount",                 # 23
    "Discount Date",                   # 24
    "Displayed Terms",                 # 25
    "Sales Representative ID",         # 26
    "Accounts Receivable Account",     # 27
    "Sales Tax ID",                    # 28
    "Invoice Note",                    # 29
    "Note Prints After Line Items",    # 30
    "Statement Note",                  # 31
    "Stmt Note Prints Before Ref",     # 32
    "Internal Note",                   # 33
    "Beginning Balance Transaction",   # 34
    "Number of Distributions",         # 35
    "Invoice/CM Distribution",         # 36
    "Apply to Invoice Distribution",   # 37
    "Apply To Sales Order",            # 38
    "Apply to Proposal",               # 39
    "Quantity",                        # 40
    "SO/Proposal Number",              # 41
    "Item ID",                         # 42
    "Serial Number",                   # 43
    "SO/Proposal Distribution",        # 44
    "Description",                     # 45
    "G/L Account",                     # 46
    "Unit Price",                      # 47
    "Tax Type",                        # 48
    "UPC / SKU",                       # 49
    "Weight",                          # 50
    "Amount",                          # 51
    "Job ID",                          # 52
    "Sales Tax Agency ID",             # 53
    "Transaction Period",              # 54
    "Transaction Number",              # 55
    "Return Authorization",            # 56
    "Voided by Transaction",           # 57
    "Recur Number",                    # 58
    "Recur Frequency",                 # 59
]

MONTHS_ES = {
    1: "enero", 2: "febrero", 3: "marzo", 4: "abril", 5: "mayo", 6: "junio",
    7: "julio", 8: "agosto", 9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre",
}
MONTH_NAME_TO_NUM = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6, "jul": 7,
    "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}

YELLOW_RGB = "FFFFFF00"


# ────────────────────────────────────────────────────────────────────────────
# Utilities
# ────────────────────────────────────────────────────────────────────────────

def is_yellow(cell) -> bool:
    """True if the cell has a solid yellow (FFFFFF00) fill."""
    fill = cell.fill
    if fill.patternType != "solid":
        return False
    fg = fill.fgColor
    if fg is None or fg.type != "rgb":
        return False
    return fg.rgb == YELLOW_RGB


def col_letter_to_idx(letter):
    """Accepts 'C' or None and returns column index or None."""
    if letter is None:
        return None
    return column_index_from_string(letter)


def to_float(v, default=0.0):
    if v is None or v == "":
        return default
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).replace(",", "").strip()
        return float(s) if s else default
    except (ValueError, TypeError):
        return default


def money(v):
    """Sage 50 money format: always 2 decimals."""
    return f"{v:.2f}"


def extract_header_fecha(sheet):
    """Find the 'Fecha:' in the top-right header of the sheet and return (year, month) if parseable."""
    for row in range(1, 10):
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=row, column=col).value
            if val and isinstance(val, str) and "fecha" in val.lower():
                # e.g., "Fecha: Marzo 15, 2026" or "Fecha: Marzo, 2026" or "Fecha: Marzo 2026"
                m = re.search(
                    r"(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)"
                    r"\D+(\d{4})",
                    val.lower(),
                )
                if m:
                    return int(m.group(2)), MONTH_NAME_TO_NUM[m.group(1)]
    return None


def parse_event_date(raw, fallback_year, fallback_month):
    """
    Parse Audico's freeform date strings into a structured representation.

    Handles:
      '28.2.26'          -> single day Feb 28, 2026
      '2-3.3'            -> range Mar 2–3, fallback year
      '2-3.3.26'         -> range Mar 2–3, 2026
      '02-03.03.26'      -> range Mar 2–3, 2026
      '14.03.26'         -> single day Mar 14, 2026
      '7.03.26'          -> single day Mar 7, 2026
      '2-4.3'            -> range Mar 2–4, fallback year
      date object        -> single day
      None / ''          -> None

    Returns dict with: start_date (date), end_date (date), is_range (bool),
                       original (str), parsed_ok (bool), note_es (str Spanish note)
    """
    result = {
        "start_date": None, "end_date": None, "is_range": False,
        "original": str(raw) if raw is not None else "", "parsed_ok": False, "note_es": "",
    }
    if raw is None or raw == "":
        return result

    # If it's already a datetime/date (Excel might parse some numerics as dates)
    if hasattr(raw, "year") and hasattr(raw, "month") and hasattr(raw, "day"):
        d = date(raw.year, raw.month, raw.day)
        result.update({
            "start_date": d, "end_date": d, "parsed_ok": True,
            "note_es": f"{d.day} de {MONTHS_ES[d.month]} de {d.year}",
        })
        return result

    s = str(raw).strip()
    if not s:
        return result

    # Normalize separators: replace hyphen-minus with - (already), keep dots
    # Pattern 1: "D-D.M.YY" or "DD-DD.MM.YY" — explicit range with year
    m = re.match(r"^\s*(\d{1,2})\s*-\s*(\d{1,2})\s*\.\s*(\d{1,2})\s*\.\s*(\d{2,4})\s*$", s)
    if m:
        d1, d2, mo, yr = int(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4))
        yr = 2000 + yr if yr < 100 else yr
        try:
            sd = date(yr, mo, d1)
            ed = date(yr, mo, d2)
            result.update({
                "start_date": sd, "end_date": ed, "is_range": True, "parsed_ok": True,
                "note_es": f"{d1} al {d2} de {MONTHS_ES[mo]} de {yr}",
            })
            return result
        except ValueError:
            pass

    # Pattern 2: "D-D.M" — range without year (use fallback)
    m = re.match(r"^\s*(\d{1,2})\s*-\s*(\d{1,2})\s*\.\s*(\d{1,2})\s*$", s)
    if m and fallback_year:
        d1, d2, mo = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            sd = date(fallback_year, mo, d1)
            ed = date(fallback_year, mo, d2)
            result.update({
                "start_date": sd, "end_date": ed, "is_range": True, "parsed_ok": True,
                "note_es": f"{d1} al {d2} de {MONTHS_ES[mo]} de {fallback_year}",
            })
            return result
        except ValueError:
            pass

    # Pattern 3: "D.M.YY" — single day with year
    m = re.match(r"^\s*(\d{1,2})\s*\.\s*(\d{1,2})\s*\.\s*(\d{2,4})\s*$", s)
    if m:
        d1, mo, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
        yr = 2000 + yr if yr < 100 else yr
        try:
            sd = date(yr, mo, d1)
            result.update({
                "start_date": sd, "end_date": sd, "parsed_ok": True,
                "note_es": f"{d1} de {MONTHS_ES[mo]} de {yr}",
            })
            return result
        except ValueError:
            pass

    # Pattern 4: "D.M" — single day without year (use fallback)
    m = re.match(r"^\s*(\d{1,2})\s*\.\s*(\d{1,2})\s*$", s)
    if m and fallback_year:
        d1, mo = int(m.group(1)), int(m.group(2))
        try:
            sd = date(fallback_year, mo, d1)
            result.update({
                "start_date": sd, "end_date": sd, "parsed_ok": True,
                "note_es": f"{d1} de {MONTHS_ES[mo]} de {fallback_year}",
            })
            return result
        except ValueError:
            pass

    # Pattern 5: "D1-D2.M.YY" where D1 > D2 = CROSS-MONTH range, e.g. "31-11.02.26" = Jan 31 → Feb 11, 2026
    # (D1 in prior month, D2 in the specified month)
    m = re.match(r"^\s*(\d{1,2})\s*-\s*(\d{1,2})\s*\.\s*(\d{1,2})\s*\.\s*(\d{2,4})\s*$", s)
    if m:
        d1, d2, mo, yr = int(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4))
        yr = 2000 + yr if yr < 100 else yr
        if d1 > d2:
            prior_mo = mo - 1 if mo > 1 else 12
            prior_yr = yr if mo > 1 else yr - 1
            try:
                sd = date(prior_yr, prior_mo, d1)
                ed = date(yr, mo, d2)
                result.update({
                    "start_date": sd, "end_date": ed, "is_range": True, "parsed_ok": True,
                    "note_es": f"{d1} de {MONTHS_ES[prior_mo]} al {d2} de {MONTHS_ES[mo]} de {yr}",
                })
                return result
            except ValueError:
                pass

    # Pattern 6: "D-M-YY" — dashes instead of dots, e.g. "25-3-26"
    m = re.match(r"^\s*(\d{1,2})\s*-\s*(\d{1,2})\s*-\s*(\d{2,4})\s*$", s)
    if m:
        d1, mo, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
        yr = 2000 + yr if yr < 100 else yr
        try:
            sd = date(yr, mo, d1)
            result.update({
                "start_date": sd, "end_date": sd, "parsed_ok": True,
                "note_es": f"{d1} de {MONTHS_ES[mo]} de {yr}",
            })
            return result
        except ValueError:
            pass

    # Pattern 7: "D1,D2.M[.YY]" — day list (non-consecutive days), e.g. "10,12.3" = Mar 10 and 12
    m = re.match(r"^\s*(\d{1,2})\s*,\s*(\d{1,2})\s*\.\s*(\d{1,2})(?:\s*\.\s*(\d{2,4}))?\s*$", s)
    if m:
        d1, d2, mo = int(m.group(1)), int(m.group(2)), int(m.group(3))
        yr = int(m.group(4)) if m.group(4) else fallback_year
        if yr and yr < 100:
            yr = 2000 + yr
        if yr:
            try:
                sd = date(yr, mo, d1)
                ed = date(yr, mo, d2)
                result.update({
                    "start_date": sd, "end_date": ed, "is_range": True, "parsed_ok": True,
                    "note_es": f"{d1} y {d2} de {MONTHS_ES[mo]} de {yr}",
                })
                return result
            except ValueError:
                pass

    # Fallback — keep original string as note
    result["note_es"] = s
    return result


# ────────────────────────────────────────────────────────────────────────────
# Core extraction
# ────────────────────────────────────────────────────────────────────────────

def extract_quotes_from_sheet(sheet, sheet_data, tab_config, fallback_year, fallback_month):
    """
    Scan the sheet and group consecutive item rows into quotes, one per yellow CLIENTE.
    `sheet` carries fill/formatting; `sheet_data` carries resolved formula values.
    Returns (quotes, review_flags) where:
      quotes = list of quote dicts (as before)
      review_flags = list of dicts with detected anomalies (discounts, notes, negative amounts)
    """
    client_col = col_letter_to_idx(tab_config["client_col"])
    desc_col = col_letter_to_idx(tab_config["desc_col"])
    qty_col = col_letter_to_idx(tab_config["qty_col"])
    days_col = col_letter_to_idx(tab_config["days_col"])
    audico_u_col = col_letter_to_idx(tab_config.get("audico_u_col"))
    total_col = col_letter_to_idx(tab_config["total_col"])
    date_col = 2  # always column B per the spec
    skip_rows = set(tab_config.get("skip_rows", []))

    quotes = []
    review_flags = []
    current_quote = None
    blank_row_streak = 0
    BLANK_ROWS_TO_CLOSE = 3
    pending_note = None  # holds the text of a "Nota:" line for the next yellow CLIENTE row

    # Spanish keywords that indicate a discount/credit/manual-handling flag
    discount_keywords = [
        "debitar", "descuento", "discount", "credito", "crédito",
        "nota de credito", "nota de crédito", "rebaja", "ajuste",
        "devolu", "reembolso", "refund",
    ]

    def _row_has_keyword(row, keyword_list):
        """Search every cell in this row for any of the keywords. Returns matched (col_letter, cell_text, keyword) or None."""
        for c in range(1, sheet.max_column + 1):
            v = sheet_data.cell(row=row, column=c).value
            if v is None:
                continue
            s = str(v).lower()
            for kw in keyword_list:
                if kw in s:
                    return (get_column_letter(c), str(v), kw)
        return None

    def _row_is_blank(row):
        """True if no content in any of the relevant columns."""
        for c in (client_col, desc_col, qty_col, total_col):
            if c is None:
                continue
            v = sheet_data.cell(row=row, column=c).value
            if v is not None and str(v).strip():
                return False
        return True

    for row in range(11, sheet.max_row + 1):
        client_cell = sheet.cell(row=row, column=client_col)
        client_val = client_cell.value
        client_val_str = str(client_val).strip() if client_val is not None else ""

        # ── Blank row tracking — close current quote after N consecutive blanks
        if _row_is_blank(row):
            blank_row_streak += 1
            if blank_row_streak >= BLANK_ROWS_TO_CLOSE and current_quote is not None:
                quotes.append(current_quote)
                current_quote = None
            continue
        else:
            blank_row_streak = 0

        # ── Discount/credit/note keyword detection (always runs, regardless of yellow status)
        kw_match = _row_has_keyword(row, discount_keywords)
        if kw_match:
            col_let, cell_text, kw = kw_match
            # Pull a useful amount value if any
            amount_val = None
            for c in [total_col, audico_u_col, col_letter_to_idx("L"), col_letter_to_idx("J")]:
                if c is None:
                    continue
                v = sheet_data.cell(row=row, column=c).value
                if isinstance(v, (int, float)) and v != 0:
                    amount_val = v
                    break

            # Spanish category label
            if kw in ("debitar",) and cell_text.lower().strip().startswith(("nota:", "nota :")):
                category = "Nota a contabilidad (memo)"
                pending_note = cell_text  # attach to next yellow CLIENTE
            elif kw in ("nota de credito", "nota de crédito", "credito", "crédito"):
                category = "Posible nota de crédito"
            elif kw in ("descuento", "discount", "rebaja"):
                category = "Posible descuento"
            elif kw in ("devolu", "reembolso", "refund"):
                category = "Posible devolución"
            elif kw in ("ajuste",):
                category = "Posible ajuste manual"
            else:
                category = "Posible descuento/crédito"

            review_flags.append({
                "tab": sheet.title,
                "fila": row,
                "columna": col_let,
                "categoria": category,
                "contenido": cell_text[:200],
                "monto": amount_val if amount_val is not None else "",
                "palabra_clave": kw,
                "accion_recomendada": "Revisar manualmente — el sistema NO incluye descuentos en las cotizaciones automáticamente.",
            })

        # ── Detect negative amount in total/L column (often = end-of-tab credit)
        l_col_idx = col_letter_to_idx("L")
        for cidx in [total_col, l_col_idx]:
            if cidx is None:
                continue
            v = sheet_data.cell(row=row, column=cidx).value
            if isinstance(v, (int, float)) and v < 0:
                # Avoid double-flagging if the keyword check already caught this row
                already = any(rf["fila"] == row and rf["tab"] == sheet.title for rf in review_flags)
                if not already:
                    # Get the description from H column (totals area uses H), or from desc_col, or whatever's nearby
                    h_col = col_letter_to_idx("H")
                    nearby = sheet_data.cell(row=row, column=h_col).value if h_col else None
                    if nearby is None:
                        nearby = sheet_data.cell(row=row, column=desc_col).value if desc_col else ""
                    review_flags.append({
                        "tab": sheet.title,
                        "fila": row,
                        "columna": get_column_letter(cidx),
                        "categoria": "Monto negativo detectado",
                        "contenido": str(nearby or "")[:200],
                        "monto": v,
                        "palabra_clave": "(monto < 0)",
                        "accion_recomendada": "Revisar — montos negativos suelen ser créditos o ajustes que se manejan aparte.",
                    })

        # A new quote starts on a yellow-highlighted CLIENTE row
        if client_val_str and is_yellow(client_cell) and client_val_str.upper() != "CLIENTE":
            # Close any previous quote
            if current_quote is not None:
                quotes.append(current_quote)
            date_raw = sheet_data.cell(row=row, column=date_col).value
            parsed = parse_event_date(date_raw, fallback_year, fallback_month)
            current_quote = {
                "client_row": row,
                "client_name": client_val_str,
                "event_date_raw": date_raw,
                "event_date": parsed,
                "items": [],
                "skipped": row in skip_rows,
                "skip_reason": "Marcada en config 'skip_rows'" if row in skip_rows else "",
                "preceded_by_note": pending_note,
            }
            # If a "Nota:" preceded this quote, also flag it for review
            if pending_note:
                review_flags.append({
                    "tab": sheet.title,
                    "fila": row,
                    "columna": tab_config["client_col"],
                    "categoria": "Cotización precedida por nota",
                    "contenido": f"Cliente: {client_val_str} — Nota previa: {pending_note[:120]}",
                    "monto": "",
                    "palabra_clave": "nota+cotización",
                    "accion_recomendada": "Verificar la nota antes de aprobar la cotización (puede requerir ajuste manual).",
                })
                pending_note = None
            # Add this row's item (first item of the quote lives on the same row as CLIENTE)
            _add_item_if_present(sheet_data, row, desc_col, qty_col, days_col, audico_u_col, total_col, current_quote)
            continue

        # A new NON-yellow client stops the current quote (different event, non-billed)
        if client_val_str and client_val_str.upper() != "CLIENTE" and not is_yellow(client_cell):
            if current_quote is not None:
                quotes.append(current_quote)
                current_quote = None
            # Reset pending note if it doesn't get used
            pending_note = None
            continue

        # Otherwise, if we're inside a quote and the row has a description, it's a continuation item
        if current_quote is not None:
            desc_val = sheet_data.cell(row=row, column=desc_col).value
            if desc_val is not None and str(desc_val).strip():
                _add_item_if_present(sheet_data, row, desc_col, qty_col, days_col, audico_u_col, total_col, current_quote)

    # Close final quote
    if current_quote is not None:
        quotes.append(current_quote)

    return quotes, review_flags


def _add_item_if_present(sheet, row, desc_col, qty_col, days_col, audico_u_col, total_col, quote):
    """Read one item row and append to the quote, if it looks like a real item."""
    desc = sheet.cell(row=row, column=desc_col).value
    if desc is None or not str(desc).strip():
        return
    desc_str = str(desc).strip()

    qty = to_float(sheet.cell(row=row, column=qty_col).value, 1.0) or 1.0
    days = to_float(sheet.cell(row=row, column=days_col).value, 1.0) or 1.0
    source_total = to_float(sheet.cell(row=row, column=total_col).value, 0.0)

    if audico_u_col is not None:
        unit_price = to_float(sheet.cell(row=row, column=audico_u_col).value, 0.0)
    else:
        # panamazing case: no AUDICO U column; derive from total / (qty * days)
        denom = qty * days
        unit_price = source_total / denom if denom else 0.0

    quote["items"].append({
        "row": row,
        "description": desc_str,
        "qty": qty,
        "days": days,
        "unit_price": unit_price,
        "computed_total": unit_price * qty * days,
        "source_total": source_total,
    })


# ────────────────────────────────────────────────────────────────────────────
# CSV row builders
# ────────────────────────────────────────────────────────────────────────────

def build_quote_rows(quote, tab_config, defaults, quote_date, good_thru_date):
    """
    Build the list of CSV rows (each a dict of header→value) that represent this quote.
    Follows the exact Peachtree pattern:
      Dist 0: ITBMS tax line (Amount = -tax, G/L 219)
      Dist 1..N: Line items (Amount = -(unit*qty*days), G/L 403, Tax 1)
      Dist N+1: '***' separator
      Dist N+2: 'Evento: <client name>'
      Dist N+3: 'Fecha: <event date range>'
    """
    items = quote["items"]
    n_items = len(items)
    total_dists = n_items + 4  # tax + N items + *** + Evento + Fecha

    subtotal = sum(it["computed_total"] for it in items)
    tax_amount = round(subtotal * defaults["sales_tax_rate"], 2)
    grand_total = round(subtotal + tax_amount, 2)

    def _peachtree_date(d: date) -> str:
        """Peachtree format: M/D/YY with no leading zeros, e.g. 3/12/26"""
        return f"{d.month}/{d.day}/{d.year % 100:02d}"

    date_str = _peachtree_date(quote_date)
    good_thru_str = _peachtree_date(good_thru_date)

    # Common header fields repeated on every row.
    # Field count and order MUST match the 59-column Sage 50 Pro import schema (see SALES_HEADER above).
    # Booleans use TRUE/FALSE strings — verified against Sage's own export.
    # The CSV writer uses extrasaction='ignore' so any extra keys here would be dropped, but we
    # only set the 59 fields Sage expects.
    common = {
        "Customer ID": tab_config["customer_id"],
        "Invoice/CM #": "",  # left blank so Sage auto-assigns on conversion
        "Apply to Invoice Number": "",
        "Credit Memo": "FALSE",
        "Progress Billing Invoice": "FALSE",
        "Date": date_str,
        "Ship By": "",
        "Quote": "TRUE",
        "Quote #": "",  # auto-assigned by Sage
        "Quote Good Thru Date": good_thru_str,
        "Drop Ship": "FALSE",
        "Ship to Name": tab_config["ship_to_name"],
        "Ship to Address-Line One": tab_config["ship_to_address"],
        "Ship to Address-Line Two": "",
        "Ship to City": tab_config["ship_to_city"],
        "Ship to State": "",
        "Ship to Zipcode": "",
        "Ship to Country": "",
        "Customer PO": "",
        "Ship Via": defaults["ship_via"],
        "Ship Date": "",
        "Date Due": good_thru_str,
        "Discount Amount": "0.00",
        "Discount Date": date_str,
        "Displayed Terms": defaults["displayed_terms"],
        "Sales Representative ID": "",
        "Accounts Receivable Account": defaults["accounts_receivable_account"],
        "Sales Tax ID": defaults["sales_tax_id"],
        "Invoice Note": "",
        "Note Prints After Line Items": "FALSE",
        "Statement Note": "",
        "Stmt Note Prints Before Ref": "FALSE",
        "Internal Note": "",
        "Beginning Balance Transaction": "FALSE",
        "Number of Distributions": str(total_dists),
        "Apply to Invoice Distribution": "0",
        "Apply To Sales Order": "FALSE",
        "Apply to Proposal": "FALSE",
        "SO/Proposal Number": "",
        "Item ID": "",
        "Serial Number": "",
        "SO/Proposal Distribution": "0",
        "UPC / SKU": "",
        "Weight": "0.00",
        "Job ID": "",
        "Transaction Period": "",
        "Transaction Number": "",
        "Return Authorization": "",
        "Voided by Transaction": "",
        "Recur Number": "0",
        "Recur Frequency": "0",
    }

    rows_out = []

    # Dist 0 — ITBMS tax line
    r0 = dict(common)
    r0.update({
        "Invoice/CM Distribution": "0",
        "Quantity": "0.00",
        "Description": "ITBMS",
        "G/L Account": defaults["gl_tax_line"],
        "Unit Price": "0.00",
        "Tax Type": "0",
        "Amount": money(-tax_amount),
        "Sales Tax Agency ID": "ITBMS",
    })
    rows_out.append(r0)

    # Dist 1..N — product lines
    for idx, item in enumerate(items, start=1):
        r = dict(common)
        r.update({
            "Invoice/CM Distribution": str(idx),
            "Quantity": money(item["qty"]),
            "Description": item["description"],
            "G/L Account": defaults["gl_product_line"],
            "Unit Price": money(item["unit_price"] * item["days"]),  # Peachtree's "Unit Price" = per-line unit × days basis
            "Tax Type": "1",
            "Amount": money(-item["computed_total"]),
            "Sales Tax Agency ID": "",
        })
        rows_out.append(r)

    # Dist N+1 — '***' separator
    r_sep = dict(common)
    r_sep.update({
        "Invoice/CM Distribution": str(n_items + 1),
        "Quantity": "0.00",
        "Description": "***",
        "G/L Account": defaults["gl_product_line"],
        "Unit Price": "0.00",
        "Tax Type": "1",
        "Amount": "0.00",
        "Sales Tax Agency ID": "",
    })
    rows_out.append(r_sep)

    # Dist N+2 — Evento: <client name>
    r_evento = dict(common)
    r_evento.update({
        "Invoice/CM Distribution": str(n_items + 2),
        "Quantity": "0.00",
        "Description": f"Evento: {quote['client_name']}",
        "G/L Account": defaults["gl_product_line"],
        "Unit Price": "0.00",
        "Tax Type": "1",
        "Amount": "0.00",
        "Sales Tax Agency ID": "",
    })
    rows_out.append(r_evento)

    # Dist N+3 — Fecha: <event dates>
    fecha_note = quote["event_date"]["note_es"] or "(sin fecha)"
    r_fecha = dict(common)
    r_fecha.update({
        "Invoice/CM Distribution": str(n_items + 3),
        "Quantity": "0.00",
        "Description": f"Fecha: {fecha_note}",
        "G/L Account": defaults["gl_product_line"],
        "Unit Price": "0.00",
        "Tax Type": "1",
        "Amount": "0.00",
        "Sales Tax Agency ID": "",
    })
    rows_out.append(r_fecha)

    return rows_out, subtotal, tax_amount, grand_total


def slugify(name, maxlen=40):
    """Filename-safe slug from a client/event name."""
    s = re.sub(r"[^A-Za-z0-9\-_ ]+", "", name).strip()
    s = re.sub(r"\s+", "_", s)
    return s[:maxlen] or "UNKNOWN"


def write_quote_csv(rows, path):
    """Write a quote to a CSV file in Sage 50 Pro import format.

    CRITICAL: NO HEADER ROW. Sage 50 Pro's Sales Journal import does not expect a header.
    Verified against Sage's own export — first row is data, not field names.
    Field order is fixed by SALES_HEADER (59 columns); Sage matches by position.
    """
    with open(path, "w", encoding="latin-1", newline="") as f:
        w = csv.DictWriter(f, fieldnames=SALES_HEADER, extrasaction="ignore", quoting=csv.QUOTE_MINIMAL)
        # NOTE: NO writeheader() call — Sage import wants pure data rows
        for r in rows:
            w.writerow(r)


# ────────────────────────────────────────────────────────────────────────────
# Main
# ────────────────────────────────────────────────────────────────────────────

def run_conversion(xlsx_path, config_path="config/hotel_mapping.json", out_dir="./out",
                    customer_list_path=None, run_date=None):
    """
    Core conversion logic as a library function.
    Returns dict with: quotes_written, not_processed (list), warnings (list),
                       summary_rows (list), out_dir (Path).
    """
    with open(config_path) as f:
        config = json.load(f)

    tabs_cfg = config["tabs"]
    ignore_tabs = set(config.get("ignore_tabs", []))
    defaults = config["defaults"]

    # Optional: load customer list for verification
    known_customer_ids = None
    if customer_list_path:
        wb_cust = load_workbook(customer_list_path, data_only=True)
        sheet_cust = wb_cust.active
        known_customer_ids = set()
        for row in range(2, sheet_cust.max_row + 1):
            cid = sheet_cust.cell(row=row, column=1).value
            if cid:
                known_customer_ids.add(str(cid).strip())

    # Run date
    if run_date is None:
        run_date = date.today()
    elif isinstance(run_date, str):
        run_date = date.fromisoformat(run_date)
    good_thru_date = run_date + timedelta(days=defaults["quote_good_thru_days"])

    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    # Load xlsx TWICE: once with formulas (for yellow fill detection) and once data_only (for computed formula values)
    wb = load_workbook(xlsx_path)
    wb_data = load_workbook(xlsx_path, data_only=True)

    # Collectors
    summary_rows = []
    not_processed_rows = []
    warnings = []
    all_review_flags = []

    total_quotes_written = 0

    for sheet_name in wb.sheetnames:
        norm = sheet_name.lower().strip()
        if norm in ignore_tabs:
            continue
        if norm not in tabs_cfg:
            warnings.append(f"Pestaña '{sheet_name}' no está en hotel_mapping.json — se omitió.")
            continue

        tab_cfg = tabs_cfg[norm]

        # Customer verification
        if known_customer_ids is not None and tab_cfg["customer_id"] not in known_customer_ids:
            warnings.append(
                f"Pestaña '{sheet_name}': el ID de cliente {tab_cfg['customer_id']} NO existe en la "
                f"lista actual de Peachtree. La importación podría fallar hasta que se cree el cliente."
            )

        sheet = wb[sheet_name]          # has fill info
        sheet_data = wb_data[sheet_name]  # has computed formula values
        header_date = extract_header_fecha(sheet)
        fallback_year = header_date[0] if header_date else run_date.year
        fallback_month = header_date[1] if header_date else run_date.month

        quotes, sheet_review_flags = extract_quotes_from_sheet(sheet, sheet_data, tab_cfg, fallback_year, fallback_month)
        all_review_flags.extend(sheet_review_flags)

        tab_out_dir = out_dir / slugify(sheet_name, 30)
        tab_out_dir.mkdir(exist_ok=True)

        for q in quotes:
            # Skip if configured
            if q["skipped"]:
                not_processed_rows.append({
                    "tab": sheet_name, "row": q["client_row"], "client": q["client_name"],
                    "event_date_raw": str(q["event_date_raw"] or ""),
                    "items": len(q["items"]),
                    "reason": q["skip_reason"],
                })
                continue

            # Skip empty quotes (yellow row but no items found)
            if not q["items"]:
                not_processed_rows.append({
                    "tab": sheet_name, "row": q["client_row"], "client": q["client_name"],
                    "event_date_raw": str(q["event_date_raw"] or ""),
                    "items": 0,
                    "reason": "Fila CLIENTE amarilla sin partidas debajo",
                })
                continue

            # Safety check: recomputed vs source totals
            for it in q["items"]:
                if it["source_total"] and abs(it["computed_total"] - it["source_total"]) > 0.01:
                    warnings.append(
                        f"{sheet_name} R{it['row']} '{q['client_name']}' — partida '{it['description']}': "
                        f"total recalculado (unit×cant×días = {it['computed_total']:.2f}) "
                        f"no coincide con columna J del archivo ({it['source_total']:.2f})."
                    )

            # Warn on unparseable dates
            if not q["event_date"]["parsed_ok"] and q["event_date_raw"]:
                warnings.append(
                    f"{sheet_name} R{q['client_row']} '{q['client_name']}' — no se pudo interpretar la fecha "
                    f"'{q['event_date_raw']}'; se incluye tal cual en la nota de la cotización."
                )

            rows, subtotal, tax, total = build_quote_rows(q, tab_cfg, defaults, run_date, good_thru_date)

            # Filename
            event_date_part = (
                q["event_date"]["start_date"].strftime("%Y-%m-%d")
                if q["event_date"]["start_date"] else "nodate"
            )
            fname = f"{slugify(sheet_name, 20)}_R{q['client_row']}_{slugify(q['client_name'], 30)}_{event_date_part}.csv"
            fpath = tab_out_dir / fname
            write_quote_csv(rows, fpath)

            summary_rows.append({
                "tab": sheet_name,
                "row": q["client_row"],
                "client": q["client_name"],
                "event_date": q["event_date"]["note_es"],
                "n_items": len(q["items"]),
                "subtotal": money(subtotal),
                "itbms": money(tax),
                "total": money(total),
                "file": str(fpath.relative_to(out_dir)),
            })
            total_quotes_written += 1

    # Write summary
    with open(out_dir / "_summary.csv", "w", encoding="utf-8", newline="") as f:
        if summary_rows:
            w = csv.DictWriter(f, fieldnames=list(summary_rows[0].keys()))
            w.writeheader()
            for r in summary_rows:
                w.writerow(r)

    # Write not-processed
    with open(out_dir / "_not_processed.csv", "w", encoding="utf-8", newline="") as f:
        if not_processed_rows:
            w = csv.DictWriter(f, fieldnames=list(not_processed_rows[0].keys()))
            w.writeheader()
            for r in not_processed_rows:
                w.writerow(r)
        else:
            f.write("tab,row,client,event_date_raw,items,reason\n")

    # Write review-needed (Spanish)
    with open(out_dir / "_revision_manual.csv", "w", encoding="utf-8", newline="") as f:
        if all_review_flags:
            w = csv.DictWriter(f, fieldnames=list(all_review_flags[0].keys()))
            w.writeheader()
            for r in all_review_flags:
                w.writerow(r)
        else:
            f.write("tab,fila,columna,categoria,contenido,monto,palabra_clave,accion_recomendada\n")

    # Write warnings
    with open(out_dir / "_warnings.txt", "w", encoding="utf-8") as f:
        if warnings:
            for w in warnings:
                f.write(w + "\n")
        else:
            f.write("Sin advertencias.\n")

    # Return structured result instead of printing
    return {
        "quotes_written": total_quotes_written,
        "summary_rows": summary_rows,
        "not_processed": not_processed_rows,
        "warnings": warnings,
        "review_flags": all_review_flags,
        "out_dir": out_dir,
    }


def main():
    """CLI entry point."""
    ap = argparse.ArgumentParser(description="Convert Audico EC xlsx to Peachtree quote CSVs.")
    ap.add_argument("xlsx", help="Path to EC_XX_YYYY.xlsx")
    ap.add_argument("--config", default="config/hotel_mapping.json", help="Hotel mapping JSON")
    ap.add_argument("--out", default="./out", help="Output directory")
    ap.add_argument("--customer-list", default=None,
                    help="Optional LISTA_DE_CLIENTES_POR_ID.xlsx for customer ID verification")
    ap.add_argument("--run-date", default=None, help="Override run date as YYYY-MM-DD (default: today)")
    args = ap.parse_args()

    result = run_conversion(
        xlsx_path=args.xlsx,
        config_path=args.config,
        out_dir=args.out,
        customer_list_path=args.customer_list,
        run_date=args.run_date,
    )

    print(f"Wrote {result['quotes_written']} quote CSVs to {result['out_dir']}")
    print(f"Summary: {result['out_dir'] / '_summary.csv'}")
    print(f"Not processed: {result['out_dir'] / '_not_processed.csv'} ({len(result['not_processed'])} entries)")
    print(f"Review needed: {result['out_dir'] / '_revision_manual.csv'} ({len(result['review_flags'])} flags)")
    print(f"Warnings: {result['out_dir'] / '_warnings.txt'} ({len(result['warnings'])} messages)")


if __name__ == "__main__":
    main()
