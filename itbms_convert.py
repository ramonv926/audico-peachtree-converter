"""
Peachtree General Ledger (cuenta ITBMS por pagar)  ->  Informe 43 (DGI)
=======================================================================

Segundo convertidor de la app. Toma el export crudo de Peachtree de la cuenta
219 "CTA POR PAGAR I.T.B.M.S" (hoja 'General Ledger') y produce el archivo del
Informe 43 de compras que se sube a la DGI a fin de mes.

Sigue las convenciones de convert.py: carga config desde config/*.json, devuelve
un dict de resultado estructurado y escribe los archivos de salida en out_dir.

Lógica (descubierta a partir de COMPRAS_MAYO_2026 + Informe43 de abril, que
reconcilian exactamente: 53 filas, B/.1504.36):

  FILTRO — qué filas del General Ledger son compras del Informe 43
    • Solo filas con DÉBITO. Los CRÉDITOS (Jrnl 'SJ') son el ITBMS *cobrado*
      sobre las VENTAS de la empresa -> no son compras, no van al informe.
    • Se excluyen retenciones (Reference/Trans empieza con 'RETEN'/'RETENCION'),
      el pago al fisco ('TESORO NACIONAL' / 'PAGOS DE IMPUESTOS' /
      'ITBMS POR PAGAR') y las filas resumen (Beginning/Current/Ending Balance).

  TRANSFORMACIÓN por fila
    ITBMS PAGADO  = Debit Amt
    MONTO BALBOAS = ITBMS / 0.07          (fórmula =J*14.2857142857143 como la plantilla)
    FECHA         = texto 'YYYYMMDD'       (¡texto!, no fecha — esto resuelve el dolor de cabeza)
    FACTURA       = Reference; en reembolsos se extrae 'FACT <n>' del detalle
    NOMBRE/RUC/DV = del detalle (reembolsos) o del MAESTRO config/itbms_vendors.json
    TIPO PERSONA  = inferido del formato de cédula/RUC (N / J / E)
    CONCEPTO      = del maestro / concepto_overrides; 1 por defecto
    COMPRAS B/S   = 1 (Locales) por defecto

  Las filas cuyo proveedor no está en el maestro (o reembolsos sin RUC en el
  texto) salen BLOQUEADAS y se listan en la hoja 'Revisar' + en el dict de
  resultado, igual que la pestaña CODE del sistema de pagos. Nada se inventa.
"""
from __future__ import annotations

import json
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

# ───────────────────────────────────────────────────────── constants
DEFAULT_ITBMS_RATE = 0.07
SUMMARY_ROWS = {"Beginning Balance", "Current Period Change", "Ending Balance"}
GL_SHEET_CANDIDATES = ("General Ledger", "General ledger", "GENERAL LEDGER")
BLOCKED_FILL = PatternFill("solid", fgColor="FFF3B0")

INFORME43_HEADERS = [
    "TIPO DE PERSONA", "RUC ", "DV ", "NOMBRE O RAZON SOCIAL", "FACTURA",
    "FECHA", "CONCEPTO", "COMPRAS DE BIENES Y SERVICIOS",
    "MONTO EN BALBOAS", "ITBMS PAGADO EN BALBOAS",
]

# ───────────────────────────────────────────────────────── patterns
RUC_RE = re.compile(r'(\d{3,}-\d{1,3}-\d{2,}|\d{6,}-\d-\d{3,}|[NEP]{1,2}-\d+-\d+)')
DV_RE = re.compile(r'DV\s*(\d{1,3})', re.I)
FACT_RE = re.compile(r'FAC?T\s*([A-Z0-9\-]+)', re.I)          # tolera el typo 'FCAT'
ITBMS_SUFFIX_RE = re.compile(r'\s*-\s*IT[MB]{2,3}S?\s*$', re.I)  # ' - ITBMS' / ' - ITMBS'


# ───────────────────────────────────────────────────────── helpers
def norm_name(s) -> str:
    """Normaliza un nombre de proveedor para emparejar: mayúsculas, sin acentos,
    sin puntuación, sin sufijos legales ni el ' - ITBMS' que pega Peachtree."""
    s = ITBMS_SUFFIX_RE.sub('', str(s or '')).strip()
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode()
    s = s.upper()
    s = re.sub(r'[.,]', ' ', s)
    s = re.sub(r'\b(S\s?A|INC|CORP|S\s?DE\s?RL)\b', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def infer_tipo(ruc) -> str:
    """N natural (cédula), E extranjero, J jurídico (empresa)."""
    r = (ruc or '').strip().upper()
    if r.startswith('E-'):
        return 'E'
    if r.startswith('N-') or r.startswith('PE-'):
        return 'N'
    if re.fullmatch(r'\d{1,2}-\d{1,4}-\d{1,5}', r):
        return 'N'
    return 'J'


def date_to_text(d) -> str:
    """A texto 'YYYYMMDD'. El General Ledger entrega datetime; si llegara texto,
    se intentan formatos comunes."""
    if isinstance(d, (datetime, date)):
        return d.strftime('%Y%m%d')
    s = str(d).strip()
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y'):
        try:
            return datetime.strptime(s[:10], fmt).strftime('%Y%m%d')
        except ValueError:
            pass
    return re.sub(r'\D', '', s)[:8]


def is_reten(ref, desc) -> bool:
    t = f"{ref or ''} {desc or ''}".upper().strip()
    return t.startswith('RETEN') or 'RETENCION' in t


def is_treasury(desc) -> bool:
    d = (desc or '').upper()
    return any(k in d for k in ('TESORO NACIONAL', 'PAGOS DE IMPUESTOS', 'ITBMS POR PAGAR'))


def is_reimbursement(ref) -> bool:
    return (ref or '').upper().startswith('REEMB')


def parse_reimbursement(desc: str):
    """De un blob de reembolso -> (nombre, ruc, dv, factura)."""
    m_ruc, m_dv, m_f = RUC_RE.search(desc), DV_RE.search(desc), FACT_RE.search(desc)
    nombre = (desc[:m_ruc.start()] if m_ruc else desc).strip()
    ruc = m_ruc.group(1) if m_ruc else ''
    dv = m_dv.group(1).zfill(2) if m_dv else ''
    fac = m_f.group(1) if m_f else ''
    return nombre, ruc, dv, fac


# ───────────────────────────────────────────────────────── vendor master
def load_vendors(config_path):
    """Carga config/itbms_vendors.json. Devuelve (master, overrides, declarant, rate).
    master: {norm_name: {ruc,dv,tipo,nombre,concepto}}. Acepta 'aliases' opcionales."""
    with open(config_path, encoding="utf-8") as f:
        cfg = json.load(f)

    master = {}
    for v in cfg.get("vendors", []):
        rec = {
            "nombre": v["nombre"].strip(),
            "ruc": str(v["ruc"]).strip(),
            "dv": str(v.get("dv", "")).strip().zfill(2) if v.get("dv") else "",
            "tipo": v.get("tipo") or infer_tipo(v["ruc"]),
            "concepto": int(v.get("concepto", 1)),
        }
        master[norm_name(v["nombre"])] = rec
        for alias in v.get("aliases", []):
            master[norm_name(alias)] = rec

    overrides = {norm_name(k): int(val) for k, val in cfg.get("concepto_overrides", {}).items()}
    decls = load_declarants(config_path, _cfg=cfg)
    declarant = decls[0]["ruc"] if decls else (cfg.get("declarant") or {}).get("ruc")
    rate = float(cfg.get("itbms_rate", DEFAULT_ITBMS_RATE))
    return master, overrides, declarant, rate


def load_declarants(config_path, _cfg=None):
    """Lista de empresas que pueden presentar (declarantes): [{ruc, dv, nombre}, ...].
    Soporta 'declarants' (lista) o el 'declarant' (singular, legado)."""
    if _cfg is None:
        with open(config_path, encoding="utf-8") as f:
            _cfg = json.load(f)
    decls = _cfg.get("declarants")
    if decls:
        return [{"ruc": str(d["ruc"]).strip(), "dv": str(d.get("dv", "")).strip(),
                 "nombre": d.get("nombre", d["ruc"])} for d in decls]
    d = _cfg.get("declarant")
    if d:
        return [{"ruc": str(d["ruc"]).strip(), "dv": str(d.get("dv", "")).strip(),
                 "nombre": d.get("nombre", d["ruc"])}]
    return []


def lookup_master(name: str, master: dict):
    """Match exacto; luego match por prefijo para sobrevivir la truncación de
    nombres de Peachtree (~30 chars), p.ej. 'MEETING & SHOW TECHNOLOGIES, S'."""
    key = norm_name(name)
    if key in master:
        return master[key]
    if len(key) >= 12:
        cands = [v for k, v in master.items() if k.startswith(key) or key.startswith(k)]
        if len(cands) == 1:
            return cands[0]
    return None


# ───────────────────────────────────────────────────────── core transform
def _gl_sheet(wb):
    for name in GL_SHEET_CANDIDATES:
        if name in wb.sheetnames:
            return wb[name]
    return wb[wb.sheetnames[0]]  # fallback: primera hoja


def transform_rows(gl_rows, master, overrides, self_rucs, rate):
    """gl_rows: tuplas (acct_id, acct_desc, date, ref, jrnl, tdesc, debit, credit, balance).
    Devuelve (lines, stats). Cada line es un dict listo para escribir / mostrar."""
    lines = []
    stats = {"sj_skipped": 0, "reten": 0, "treasury": 0, "summary": 0}

    for r in gl_rows:
        _, _, dt, ref, jrnl, tdesc, debit, credit, _ = r
        if tdesc in SUMMARY_ROWS:
            stats["summary"] += 1; continue
        if debit is None:
            stats["sj_skipped"] += 1; continue          # crédito = ITBMS de ventas
        if is_reten(ref, tdesc):
            stats["reten"] += 1; continue
        if is_treasury(tdesc):
            stats["treasury"] += 1; continue

        itbms = round(float(debit), 2)
        ln = {
            "tipo": "", "ruc": "", "dv": "", "nombre": "", "factura": "",
            "fecha": date_to_text(dt), "concepto": 1, "compras_bs": 1,
            "monto": round(itbms / rate, 2), "itbms": itbms,
            "source": str(tdesc), "blocked": False, "reasons": [],
        }

        if is_reimbursement(ref):
            nombre, ruc, dv, fac = parse_reimbursement(str(tdesc))
            ln.update(nombre=nombre, ruc=ruc, dv=dv, factura=fac)
            if not ruc:
                ln["blocked"] = True
                ln["reasons"].append("Reembolso sin RUC en el detalle — completar manualmente")
            else:
                mm = lookup_master(nombre, master)
                if mm and not dv:
                    ln["dv"] = mm["dv"]
        else:
            nombre = ITBMS_SUFFIX_RE.sub('', str(tdesc)).strip()
            ln["nombre"] = nombre
            ln["factura"] = '' if ref is None else str(ref).strip()
            m = lookup_master(nombre, master)
            if m:
                ln.update(ruc=m["ruc"], dv=m["dv"], nombre=m["nombre"])
            else:
                ln["blocked"] = True
                ln["reasons"].append(f"Proveedor sin RUC/DV en el maestro: '{nombre}'")

        # sanity: el RUC del proveedor no puede ser el de una empresa del grupo (declarante)
        if self_rucs and ln["ruc"] in self_rucs:
            ln["blocked"] = True
            ln["reasons"].append(f"RUC = empresa del grupo/declarante ({ln['ruc']}); error de captura, revisar")

        ln["tipo"] = infer_tipo(ln["ruc"]) if ln["ruc"] else ""
        ln["concepto"] = overrides.get(norm_name(ln["nombre"]),
                                       (lookup_master(ln["nombre"], master) or {}).get("concepto", 1))
        lines.append(ln)

    return lines, stats


# ───────────────────────────────────────────────────────── output writer
def write_informe43(lines, out_path: Path):
    """Escribe el .xlsx en formato Informe 43. Columnas A-H como TEXTO (@),
    FECHA como texto 'YYYYMMDD', MONTO como fórmula =J*14.2857142857143.
    Filas bloqueadas resaltadas en amarillo + hoja 'Revisar'."""
    out = Workbook()
    ws = out.active
    ws.title = "Hoja1"
    ws["A1"] = "INFORME 43 - FORMATO A DILIGENCIAR"
    ws["A2"] = "Esta sección de encabezado no se debe modificar. "
    ws["A3"] = "Los datos de este informe deben ser registrados a partir de la línea 5 en adelante."
    for c, h in enumerate(INFORME43_HEADERS, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = Font(bold=True)

    row = 5
    for ln in lines:
        fill = BLOCKED_FILL if ln["blocked"] else None
        text_vals = [ln["tipo"], ln["ruc"], ln["dv"], ln["nombre"], str(ln["factura"]),
                     ln["fecha"], ln["concepto"], ln["compras_bs"]]
        for c, v in enumerate(text_vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.number_format = '@'                      # TEXTO — clave para la fecha
            if fill:
                cell.fill = fill
        m = ws.cell(row=row, column=9, value=f"=+J{row}*14.2857142857143")
        m.number_format = "0.00"
        j = ws.cell(row=row, column=10, value=ln["itbms"])
        j.number_format = "0.00"
        if fill:
            m.fill = fill; j.fill = fill
        row += 1

    blocked = [ln for ln in lines if ln["blocked"]]
    qa = out.create_sheet("Revisar")
    qa.append(["FECHA", "NOMBRE", "FACTURA", "ITBMS", "MOTIVO"])
    for c in range(1, 6):
        qa.cell(row=1, column=c).font = Font(bold=True)
    for ln in blocked:
        qa.append([ln["fecha"], ln["nombre"], str(ln["factura"]), ln["itbms"], "; ".join(ln["reasons"])])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out.save(out_path)
    return out_path


# ───────────────────────────────────────────────────────── public entry
def detect_periods(xlsx_path):
    """Lee el Mayor y devuelve un Counter {YYYYMM: n_filas_de_compra} de los
    meses presentes (solo filas que SÍ son compras del Informe 43). Permite al
    front-end mostrar qué meses contiene el archivo antes de convertir."""
    from collections import Counter
    wb = load_workbook(xlsx_path, data_only=True)
    ws = _gl_sheet(wb)
    rows = list(ws.iter_rows(values_only=True))[1:]
    wb.close()
    months = Counter()
    for r in rows:
        _, _, dt, ref, jrnl, tdesc, debit, credit, _ = r
        if tdesc in SUMMARY_ROWS or debit is None:
            continue
        if is_reten(ref, tdesc) or is_treasury(tdesc):
            continue
        ym = date_to_text(dt)[:6]
        if ym:
            months[ym] += 1
    return months


# ───────────────────────────────────────────────────────── editable-grid support
EDITOR_FIELDS = ["estado", "fecha", "nombre", "factura", "tipo", "ruc", "dv",
                 "concepto", "itbms", "monto"]


# Campos OBLIGATORIOS del Informe 43 (Fecha/ITBMS/Monto vienen del Mayor)
REQUIRED_LABELS = [
    ("tipo", "Tipo"), ("ruc", "RUC/Cédula"), ("dv", "DV"),
    ("nombre", "Nombre"), ("factura", "Factura"), ("concepto", "Concepto"),
]


def _as_ruc_set(declarant_rucs):
    """Acepta None, str, o coleccion -> set de RUCs 'propios' (del grupo)."""
    if declarant_rucs is None:
        return set()
    if isinstance(declarant_rucs, str):
        return {declarant_rucs}
    return set(declarant_rucs)


def detect_declarant_from_gl(gl_rows, declarants):
    """Detecta que empresa presenta, buscando su RUC en la fila de pago al fisco
    (Tesoro Nacional) del Mayor. Devuelve el RUC detectado o None."""
    import re as _re
    for r in gl_rows:
        tdesc = r[5] if len(r) > 5 else None
        if not is_treasury(tdesc):
            continue
        digits = _re.sub(r"\\D", "", str(tdesc))
        for d in declarants:
            principal = _re.sub(r"\\D", "", str(d["ruc"]).split("-")[0])
            if principal and principal in digits:
                return d["ruc"]
    return None


def missing_fields(rec, declarant_rucs=None):
    """Lista de campos obligatorios que faltan en una fila (vacía = fila OK).
    declarant_rucs: str o coleccion de RUCs propios del grupo (AUDICO, 3S)."""
    self_rucs = _as_ruc_set(declarant_rucs)
    miss = []
    for key, label in REQUIRED_LABELS:
        v = rec.get(key)
        if v is None or str(v).strip() == "":
            miss.append(label)
    ruc = str(rec.get("ruc") or "").strip()
    if ruc and ruc in self_rucs:
        miss.append("RUC = declarante (corregir)")
    return miss


def _row_status(rec, declarant_rucs=None):
    """Texto de estado para el grid editable."""
    miss = missing_fields(rec, declarant_rucs)
    return "OK" if not miss else "Falta: " + ", ".join(miss)


def lines_to_editor_rows(lines, rate=DEFAULT_ITBMS_RATE, declarant_rucs=None):
    """Aplana las líneas internas a filas planas para st.data_editor."""
    rows = []
    for ln in lines:
        rec = {
            "fecha": ln["fecha"], "nombre": ln["nombre"], "factura": str(ln["factura"]),
            "tipo": ln["tipo"], "ruc": ln["ruc"], "dv": ln["dv"],
            "concepto": int(ln["concepto"]), "itbms": float(ln["itbms"]),
            "monto": round(float(ln["itbms"]) / rate, 2),
        }
        rec["estado"] = _row_status(rec, declarant_rucs)
        rows.append(rec)
    return rows


def finalize_records(records, declarant_rucs=None, rate=DEFAULT_ITBMS_RATE):
    """Toma las filas editadas (lista de dicts) y devuelve (lines, pending).
    Normaliza dv, infiere tipo si falta, recalcula monto y marca bloqueadas."""
    lines, pending = [], 0
    for rec in records:
        ruc = str(rec.get("ruc") or "").strip()
        dv = str(rec.get("dv") or "").strip()
        if dv and dv.isdigit():
            dv = dv.zfill(2)
        tipo = str(rec.get("tipo") or "").strip().upper()   # obligatorio: no se infiere en silencio
        itbms = round(float(rec.get("itbms") or 0), 2)
        ln = {
            "tipo": tipo, "ruc": ruc, "dv": dv,
            "nombre": str(rec.get("nombre") or "").strip(),
            "factura": str(rec.get("factura") or "").strip(),
            "fecha": str(rec.get("fecha") or "").strip(),
            "concepto": int(rec.get("concepto") or 1), "compras_bs": 1,
            "monto": round(itbms / rate, 2), "itbms": itbms,
            "blocked": False, "reasons": [],
        }
        miss = missing_fields(rec, declarant_rucs)
        if miss:
            ln["blocked"] = True
            ln["reasons"].append("Falta: " + ", ".join(miss))
            pending += 1
        lines.append(ln)
    return lines, pending


def vendor_master_delta(lines, config_path):
    """Devuelve el dict de config con los proveedores NUEVOS ya completados
    agregados al maestro (para que el usuario lo descargue y tú lo commitees).
    No escribe a disco: en Streamlit Cloud el FS es efímero."""
    with open(config_path, encoding="utf-8") as f:
        cfg = json.load(f)
    existing = {norm_name(v["nombre"]) for v in cfg.get("vendors", [])}
    seen, added = set(existing), 0
    for ln in lines:
        if ln["blocked"] or not ln["ruc"]:
            continue
        key = norm_name(ln["nombre"])
        if key in seen:
            continue
        cfg.setdefault("vendors", []).append({
            "nombre": ln["nombre"], "ruc": ln["ruc"], "dv": ln["dv"],
            "tipo": ln["tipo"], "concepto": int(ln["concepto"]),
        })
        seen.add(key); added += 1
    cfg["vendors"] = sorted(cfg.get("vendors", []), key=lambda v: v["nombre"].upper())
    return cfg, added


def run_itbms_conversion(xlsx_path, config_path="config/itbms_vendors.json",
                         out_dir="./out", period=None, filter_to_period=True,
                         declarant_ruc=None):
    """Convierte el General Ledger crudo de Peachtree al Informe 43.
    Devuelve un dict de resultado estructurado (mismo estilo que convert.py)."""
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    master, overrides, cfg_declarant, rate = load_vendors(config_path)
    declarants = load_declarants(config_path)
    self_rucs = {d["ruc"] for d in declarants} or ({cfg_declarant} if cfg_declarant else set())

    wb = load_workbook(xlsx_path, data_only=True)
    ws = _gl_sheet(wb)
    gl_rows = list(ws.iter_rows(values_only=True))[1:]   # quita encabezado
    wb.close()

    # empresa para el NOMBRE del archivo: override explicito > detectada del Mayor > primera configurada
    file_declarant = (declarant_ruc
                      or detect_declarant_from_gl(gl_rows, declarants)
                      or (declarants[0]["ruc"] if declarants else cfg_declarant))

    lines, stats = transform_rows(gl_rows, master, overrides, self_rucs, rate)

    # meses presentes en el archivo
    from collections import Counter
    months_present = Counter(ln["fecha"][:6] for ln in lines if ln["fecha"])

    # período: el indicado, o el mes dominante del archivo
    if not period:
        period = months_present.most_common(1)[0][0] if months_present else datetime.today().strftime("%Y%m")

    # filtrar a un solo mes (el Informe 43 es estrictamente mensual)
    dropped = []
    if filter_to_period and period:
        dropped = [ln for ln in lines if ln["fecha"][:6] != period]
        lines = [ln for ln in lines if ln["fecha"][:6] == period]

    decl = (file_declarant or "DECLARANTE")
    out_name = f"Informe43_{decl}_{period}.xlsx"
    out_path = write_informe43(lines, out_dir / out_name)

    resolved = [ln for ln in lines if not ln["blocked"]]
    blocked = [ln for ln in lines if ln["blocked"]]

    summary_rows = [{
        "tipo": ln["tipo"], "ruc": ln["ruc"], "dv": ln["dv"], "nombre": ln["nombre"],
        "factura": str(ln["factura"]), "fecha": ln["fecha"], "concepto": ln["concepto"],
        "monto": ln["monto"], "itbms": ln["itbms"],
        "estado": "BLOQUEADA" if ln["blocked"] else "OK",
    } for ln in lines]

    not_processed = [{
        "fecha": ln["fecha"], "nombre": ln["nombre"], "factura": str(ln["factura"]),
        "itbms": ln["itbms"], "motivo": "; ".join(ln["reasons"]),
    } for ln in blocked]

    warnings = []
    if stats["sj_skipped"]:
        warnings.append(f"Se omitieron {stats['sj_skipped']} filas de VENTAS (Jrnl SJ / crédito) — "
                        "esas son ITBMS cobrado, no compras.")
    if stats["reten"]:
        warnings.append(f"Se omitieron {stats['reten']} retenciones (RETEN/RETENCION).")
    if stats["treasury"]:
        warnings.append(f"Se omitió {stats['treasury']} pago(s) al fisco (Tesoro Nacional).")
    if blocked:
        warnings.append(f"{len(blocked)} filas quedaron BLOQUEADAS — revisa la pestaña 'Revisión manual' "
                        "y completa el proveedor en config/itbms_vendors.json.")
    if dropped:
        otros = ", ".join(f"{m} ({n})" for m, n in Counter(d["fecha"][:6] for d in dropped).items())
        warnings.append(f"Se excluyeron {len(dropped)} filas de OTROS meses (no van en el informe de "
                        f"{period}): {otros}. Verifica que exportaste el rango correcto en Peachtree.")
    if not lines:
        warnings.append(f"⚠️ NINGUNA fila de compra en {period}. ¿Exportaste el mes correcto del Mayor?")

    return {
        "editor_rows": lines_to_editor_rows(lines, rate, self_rucs),
        "declarant_ruc": file_declarant,
        "self_rucs": sorted(self_rucs),
        "declarant_detected": detect_declarant_from_gl(gl_rows, declarants) is not None,
        "rows_written": len(resolved),
        "rows_blocked": len(blocked),
        "summary_rows": summary_rows,
        "not_processed": not_processed,
        "review_flags": [],
        "warnings": warnings,
        "totals": {
            "itbms": round(sum(ln["itbms"] for ln in lines), 2),
            "monto": round(sum(ln["monto"] for ln in lines), 2),
        },
        "stats": stats,
        "period": period,
        "out_path": str(out_path),
        "out_dir": str(out_dir),
    }
